#!/usr/bin/env python3
"""Create a manual tagging template for map tile faction marks.

The stable faction mark is the rightmost icon in the bottom-right corner.
If another icon appears to its left, that left icon is only a candidate:
it can be a second faction mark, a non-faction icon, or something uncertain.
"""

from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
MAP_DATA = ROOT / "map" / "map-data.js"
OUT_DIR = ROOT / "tools" / "map-faction-tag-template"
CROP_DIR = OUT_DIR / "crops"
PAGES_DIR = OUT_DIR / "pages"
OUT_CSV = OUT_DIR / "map-faction-tags-template.csv"
OUT_XLSX = OUT_DIR / "map-faction-tags-template.xlsx"
OUT_HTML = OUT_DIR / "index.html"
FONT_PATH = "C:/Windows/Fonts/msyh.ttc"


FACTIONS_BY_CYCLE = {
    "c1": ["minoians", "labyrinthians", "hornsworn"],
    "c2": ["helots", "cyclopes", "symmachy"],
    "c3": ["sunheirs", "delphians", "twilightWatch"],
}

FIELDS = [
    "cycle",
    "tile_id",
    "label",
    "crop_path",
    "right_faction",
    "left_icon_type",
    "left_faction_if_any",
    "notes",
]

LEFT_ICON_TYPES = ["none", "faction", "non_faction", "unsure"]


def load_tiles() -> list[dict[str, str]]:
    text = MAP_DATA.read_text(encoding="utf-8")
    match = re.search(r"window\.ATO_MAP_DATA\s*=\s*(\{.*\});?\s*$", text, re.S)
    if not match:
        raise RuntimeError("Could not parse map-data.js")
    data = json.loads(match.group(1))
    rows: list[dict[str, str]] = []
    for cycle in data["cycles"]:
        for tile in cycle["tiles"]:
            rows.append(
                {
                    "cycle": cycle["id"],
                    "tile_id": tile["id"],
                    "label": tile.get("label", tile["id"]),
                    "front": str((ROOT / "map" / tile["front"]).resolve()),
                }
            )
    return rows


def draw_candidate_guides(crop: Image.Image) -> None:
    """Draw fixed inner/outer slots based on the user's reference image."""
    draw = ImageDraw.Draw(crop)
    width, height = crop.size

    # Coordinates are normalized to the generated bottom-right context crop.
    # They match the provided reference: inner is the left candidate slot,
    # outer is the stable right faction slot.
    left_box = (
        round(width * 0.476),
        round(height * 0.625),
        round(width * 0.695),
        round(height * 0.902),
    )
    right_box = (
        round(width * 0.695),
        round(height * 0.625),
        round(width * 0.935),
        round(height * 0.902),
    )
    draw.rectangle(left_box, outline=(255, 214, 64), width=5)
    draw.rectangle(right_box, outline=(239, 83, 80), width=5)


def crop_right_bottom(tile: dict[str, str]) -> Path:
    image = Image.open(tile["front"]).convert("RGB")
    width, height = image.size

    # Wide bottom-right context. This keeps both side-by-side icons visible
    # without treating the left one as a faction automatically.
    crop = image.crop((round(width * 0.50), round(height * 0.56), width, height))
    target_w = 520
    target_h = round(crop.height * (target_w / crop.width))
    crop = crop.resize((target_w, target_h), Image.Resampling.LANCZOS)

    draw = ImageDraw.Draw(crop)
    font = ImageFont.truetype(FONT_PATH, 24)
    draw.rectangle((0, 0, target_w - 1, 36), fill=(255, 255, 255))
    draw.text((10, 4), f"{tile['cycle']}:{tile['tile_id']}", font=font, fill=(38, 50, 56))
    draw_candidate_guides(crop)

    path = CROP_DIR / f"{tile['cycle']}-{tile['tile_id']}.png"
    crop.save(path)
    return path


def make_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for tile in load_tiles():
        crop_path = crop_right_bottom(tile)
        rows.append(
            {
                "cycle": tile["cycle"],
                "tile_id": tile["tile_id"],
                "label": tile["label"],
                "crop_path": str(crop_path.resolve()),
                "right_faction": "",
                "left_icon_type": "",
                "left_faction_if_any": "",
                "notes": "",
            }
        )
    return rows


def write_csv(rows: list[dict[str, str]]) -> None:
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "map_faction_tags"
    options = wb.create_sheet("options")
    options.sheet_state = "hidden"

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for idx, field in enumerate(FIELDS + ["preview"], start=1):
        cell = ws.cell(row=1, column=idx, value=field)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(FIELDS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[field])

        img = XlsxImage(row["crop_path"])
        scale = 320 / img.width
        img.width = 320
        img.height = round(img.height * scale)
        ws.add_image(img, f"I{row_idx}")
        ws.row_dimensions[row_idx].height = 178

    option_row = 1
    all_factions = sorted({f for factions in FACTIONS_BY_CYCLE.values() for f in factions})
    for value in [""] + all_factions:
        options.cell(row=option_row, column=1, value=value)
        option_row += 1
    for idx, value in enumerate([""] + LEFT_ICON_TYPES, start=1):
        options.cell(row=idx, column=2, value=value)

    faction_dv = DataValidation(
        type="list",
        formula1=f"options!$A$1:$A${len(all_factions) + 1}",
        allow_blank=True,
    )
    ws.add_data_validation(faction_dv)
    faction_dv.add(f"E2:E{len(rows) + 1}")
    faction_dv.add(f"G2:G{len(rows) + 1}")

    type_dv = DataValidation(type="list", formula1=f"options!$B$1:$B${len(LEFT_ICON_TYPES) + 1}", allow_blank=True)
    ws.add_data_validation(type_dv)
    type_dv.add(f"F2:F{len(rows) + 1}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"
    widths = [9, 9, 10, 56, 18, 18, 20, 36, 48]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.sheet_view.showGridLines = False

    try:
        wb.save(OUT_XLSX)
        return OUT_XLSX
    except OSError:
        fallback = OUT_DIR / f"map-faction-tags-template-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
        wb.save(fallback)
        return fallback


def write_pages(rows: list[dict[str, str]]) -> None:
    cols = 2
    cell_w, cell_h = 560, 325
    per_page = 10
    font = ImageFont.truetype(FONT_PATH, 24)
    small_font = ImageFont.truetype(FONT_PATH, 18)
    page_paths: list[Path] = []

    for page_index, start in enumerate(range(0, len(rows), per_page), start=1):
        chunk = rows[start : start + per_page]
        page_rows = (len(chunk) + cols - 1) // cols
        sheet = Image.new("RGB", (cols * cell_w, page_rows * cell_h), "#f7f8f5")
        draw = ImageDraw.Draw(sheet)
        for idx, row in enumerate(chunk):
            x = (idx % cols) * cell_w
            y = (idx // cols) * cell_h
            crop = Image.open(row["crop_path"]).convert("RGB").resize((520, 245), Image.Resampling.LANCZOS)
            sheet.paste(crop, (x + 20, y + 14))
            draw.text((x + 22, y + 268), f"{row['cycle']}:{row['tile_id']}", font=font, fill="#263238")
            draw.text((x + 116, y + 274), "yellow=left candidate, red=right faction", font=small_font, fill="#546e7a")
        path = PAGES_DIR / f"map-faction-tag-page-{page_index:02d}.png"
        sheet.save(path)
        page_paths.append(path)

    sections = "\n".join(
        f'<section><h2>Page {idx}</h2><img src="pages/{path.name}" alt="page {idx}"></section>'
        for idx, path in enumerate(page_paths, start=1)
    )
    OUT_HTML.write_text(
        f"""<!doctype html>
<html lang="zh-CN">
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Map faction tag template</title>
<style>
body {{ margin:0; background:#f7f8f5; font-family:system-ui,sans-serif; color:#263238; }}
header {{ position:sticky; top:0; background:rgba(247,248,245,.96); padding:12px 14px; border-bottom:1px solid #d8dedf; }}
h1 {{ margin:0; font-size:20px; }}
h2 {{ margin:18px 12px 8px; font-size:16px; }}
img {{ display:block; width:100%; height:auto; }}
</style>
<header><h1>Map faction tag template</h1></header>
{sections}
</html>
""",
        encoding="utf-8",
    )


def main() -> int:
    CROP_DIR.mkdir(parents=True, exist_ok=True)
    PAGES_DIR.mkdir(parents=True, exist_ok=True)
    rows = make_rows()
    write_csv(rows)
    xlsx_path = write_xlsx(rows)
    write_pages(rows)
    print(xlsx_path)
    print(OUT_CSV)
    print(OUT_HTML)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
