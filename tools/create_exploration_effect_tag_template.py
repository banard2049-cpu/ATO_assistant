#!/usr/bin/env python3
"""Create a human tagging template for exploration-card effects."""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OCR_PATH = ROOT / "tools" / "ocr-exploration-all.json"
REPORT_CSV = ROOT / "tools" / "exploration-effect-automation-report.csv"
OUT_CSV = ROOT / "tools" / "exploration-effect-tags-template.csv"
OUT_XLSX = ROOT / "tools" / "exploration-effect-tags-template.xlsx"
OUT_MD = ROOT / "tools" / "exploration-effect-tags-template.md"
INDEX_PATH = ROOT / "index.html"


FIELDS = [
    "key",
    "cycle",
    "card_id",
    "name",
    "image_path",
    "automation_bucket",
    "base_effect_type",
    "base_amount",
    "base_target",
    "base_resource_key",
    "base_resource_name",
    "base_raw_text",
    "conditional_effects",
    "requires_choice",
    "choice_prompt",
    "requires_state_check",
    "state_key",
    "timing",
    "can_apply_automatically",
    "confidence",
    "needs_icon_review",
    "notes",
    "ocr_full",
]


EFFECT_TYPES = [
    "",
    "gain_resource",
    "lose_resource",
    "gain_track",
    "lose_track",
    "gain_titan",
    "lose_titan",
    "gain_priest",
    "gain_rare_resource",
    "gain_argonaut_stat",
    "add_timeline_event",
    "start_battle",
    "resolve_story_event",
    "map_change",
    "other",
]

TARGETS = [
    "",
    "argo_resources",
    "argo_track",
    "argo_knowledge",
    "inward_odyssey",
    "chosen_argonaut",
    "each_argonaut",
    "group",
    "timeline",
    "map",
    "story",
    "battle",
    "deck",
]

BOOLS = ["", "yes", "no"]
CONFIDENCE = ["", "high", "medium", "low"]
TIMINGS = ["", "on_draw", "end_of_exploration_step", "on_settle", "until_next_timeline_battle", "permanent", "other"]


def parse_catalog() -> dict[str, dict[str, str]]:
    text = INDEX_PATH.read_text(encoding="utf-8")
    catalog: dict[str, dict[str, str]] = {}
    cycle = None
    for line in text.splitlines():
        cycle_match = re.match(r"\s*(c[123]):\s*\[", line)
        if cycle_match:
            cycle = cycle_match.group(1)
            continue
        if cycle and re.match(r"\s*const explorationCornerNumbers", line):
            break
        entry_match = re.search(r'\{\s*id:\s*"(\d+)",\s*name:\s*"([^"]+)"\s*\}', line)
        if cycle and entry_match:
            card_id, name = entry_match.groups()
            catalog[f"{cycle}:{card_id}"] = {"cycle": cycle, "card_id": card_id, "name": name}
        range_match = re.search(r"Array\.from\(\{\s*length:\s*(\d+)\s*\}.*String\((\d+)\s*\+\s*index\)", line)
        if cycle and range_match:
            count, start = map(int, range_match.groups())
            for card_id in range(start, start + count):
                catalog[f"{cycle}:{card_id}"] = {
                    "cycle": cycle,
                    "card_id": str(card_id),
                    "name": f"Cycle III 探索卡 {card_id}",
                }
    return catalog


def key_from_file(value: str) -> str:
    path = Path(value)
    return f"{path.parts[-2]}:{path.stem}"


def load_report() -> dict[str, dict[str, str]]:
    if not REPORT_CSV.exists():
        return {}
    with REPORT_CSV.open(encoding="utf-8-sig", newline="") as handle:
        return {row["key"]: row for row in csv.DictReader(handle)}


def guess_base_effect(ocr_full: str) -> dict[str, str]:
    text = re.sub(r"\s+", " ", ocr_full)
    match = re.search(r"\bGain\s+(\d+)\s+([A-Za-z][A-Za-z \-]+?)\s+resources?\b", text, re.I)
    if match:
        return {
            "base_effect_type": "gain_resource",
            "base_amount": match.group(1),
            "base_resource_name": match.group(2).strip(),
            "base_target": "argo_resources",
            "base_raw_text": match.group(0),
            "can_apply_automatically": "yes",
            "confidence": "medium",
        }
    titan = re.search(r"\bGain\s*\+?\s*(\d+)\s+Titan\b", text, re.I)
    if titan:
        return {
            "base_effect_type": "gain_titan",
            "base_amount": titan.group(1),
            "base_target": "argo_track",
            "base_raw_text": titan.group(0),
            "can_apply_automatically": "no",
            "confidence": "medium",
        }
    return {
        "base_effect_type": "",
        "base_amount": "",
        "base_target": "",
        "base_resource_name": "",
        "base_raw_text": "",
        "can_apply_automatically": "no",
        "confidence": "",
    }


def build_rows() -> list[dict[str, str]]:
    catalog = parse_catalog()
    report = load_report()
    ocr = json.loads(OCR_PATH.read_text(encoding="utf-8"))
    rows = []
    for card in ocr["cards"]:
        key = key_from_file(card["file"])
        meta = catalog.get(key, {})
        ocr_full = str(card.get("text", {}).get("full", "")).replace("\n", " / ")
        guess = guess_base_effect(ocr_full)
        bucket = report.get(key, {}).get("status", "")
        row = {field: "" for field in FIELDS}
        row.update({
            "key": key,
            "cycle": meta.get("cycle", key.split(":")[0]),
            "card_id": meta.get("card_id", key.split(":")[1]),
            "name": meta.get("name", ""),
            "image_path": str((ROOT / card["file"]).resolve()),
            "automation_bucket": bucket,
            "timing": "on_settle",
            "requires_choice": "no" if bucket == "base-auto" else "",
            "requires_state_check": "yes" if bucket in {"base-auto", "manual", "review"} else "",
            "needs_icon_review": "yes",
            "ocr_full": ocr_full,
        })
        row.update(guess)
        rows.append(row)
    rows.sort(key=lambda item: (item["cycle"], int(item["card_id"])))
    return rows


def write_csv(rows: list[dict[str, str]]) -> None:
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def add_validation(ws, column_name: str, options: list[str], start_row: int, end_row: int) -> None:
    col = FIELDS.index(column_name) + 1
    hidden = ws.parent["选项"]
    option_col = hidden.max_column + 1
    hidden.cell(row=1, column=option_col, value=column_name)
    for idx, option in enumerate(options, start=2):
        hidden.cell(row=idx, column=option_col, value=option)
    col_letter = hidden.cell(row=1, column=option_col).column_letter
    formula = f"选项!${col_letter}$2:${col_letter}${len(options) + 1}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{ws.cell(row=1, column=col).column_letter}{start_row}:{ws.cell(row=1, column=col).column_letter}{end_row}")


def write_xlsx(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "探索卡效果打标"
    options_ws = wb.create_sheet("选项")
    options_ws.sheet_state = "hidden"

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for col_idx, field in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(FIELDS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=field in {"conditional_effects", "notes", "ocr_full", "base_raw_text", "choice_prompt"})

    widths = {
        "key": 12,
        "cycle": 8,
        "card_id": 10,
        "name": 28,
        "image_path": 42,
        "automation_bucket": 18,
        "base_effect_type": 20,
        "base_amount": 12,
        "base_target": 20,
        "base_resource_key": 22,
        "base_resource_name": 24,
        "base_raw_text": 36,
        "conditional_effects": 54,
        "requires_choice": 15,
        "choice_prompt": 36,
        "requires_state_check": 18,
        "state_key": 20,
        "timing": 26,
        "can_apply_automatically": 22,
        "confidence": 12,
        "needs_icon_review": 18,
        "notes": 44,
        "ocr_full": 72,
    }
    for idx, field in enumerate(FIELDS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths.get(field, 18)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    end_row = len(rows) + 1
    add_validation(ws, "base_effect_type", EFFECT_TYPES, 2, end_row)
    add_validation(ws, "base_target", TARGETS, 2, end_row)
    add_validation(ws, "requires_choice", BOOLS, 2, end_row)
    add_validation(ws, "requires_state_check", BOOLS, 2, end_row)
    add_validation(ws, "can_apply_automatically", BOOLS, 2, end_row)
    add_validation(ws, "confidence", CONFIDENCE, 2, end_row)
    add_validation(ws, "needs_icon_review", BOOLS, 2, end_row)
    add_validation(ws, "timing", TIMINGS, 2, end_row)

    ws.row_dimensions[1].height = 36
    for row_idx in range(2, min(end_row, 30) + 1):
        ws.row_dimensions[row_idx].height = 76
    wb.save(OUT_XLSX)


def write_docs() -> None:
    OUT_MD.write_text(
        """# 探索卡效果打标模板说明

推荐编辑 `exploration-effect-tags-template.xlsx`。CSV 版用于脚本读取。

## 填法

- `base_effect_type`：卡牌无条件基础效果，例如 `gain_resource`。
- `base_amount`：基础效果数量，只填数字。
- `base_target`：效果作用对象，例如 `argo_resources`、`chosen_argonaut`、`group`。
- `base_resource_key`：后续程序用的资源 key。可以先空着，等统一资源映射时补。
- `base_resource_name`：卡面英文资源名，OCR 已尽量预填。
- `base_raw_text`：对应卡面原文，方便复核。
- `conditional_effects`：把“若 xxx，额外获取/改为获取/失去/触发事件”写成结构化短句。建议一行一个条件。
- `requires_choice`：是否需要玩家选择，例如 chosen Argonaut、or 二选一。
- `choice_prompt`：如果需要选择，写给界面显示的问题。
- `requires_state_check`：是否需要程序检查状态，例如外交、地图图标、资源数量、Inward Odyssey。
- `state_key`：状态名，例如 `diplomacy.minoians`、`currentTile.icon`、`resources.violentAmbrosia`。
- `timing`：触发时机。大多数探索卡基础收益可用 `on_settle`。
- `can_apply_automatically`：是否能在确认后直接改记录表。
- `confidence`：你复核后的置信度。
- `needs_icon_review`：是否还需要看图标确认条件。
- `notes`：任何人话备注。

## 建议约定

基础资源收益尽量拆出来，例如：

`Gain 2 Trireme resources. 若 Friendly，Gain 3 instead.`

可以填：

- `base_effect_type = gain_resource`
- `base_amount = 2`
- `base_target = argo_resources`
- `base_resource_name = Trireme`
- `conditional_effects = if diplomacy friendly: replace base amount with 3`
- `requires_state_check = yes`
- `state_key = diplomacy.<faction>`

这样后面程序可以先自动准备基础收益，再根据条件决定是否替换或弹确认。
""",
        encoding="utf-8",
    )


def main() -> int:
    rows = build_rows()
    write_csv(rows)
    write_xlsx(rows)
    write_docs()
    print(OUT_XLSX)
    print(OUT_CSV)
    print(OUT_MD)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
