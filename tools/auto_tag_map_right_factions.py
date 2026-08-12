#!/usr/bin/env python3
"""Auto-tag map tile right-side factions with front/back agreement.

Pipeline:
1. Match front outer slot and back top badge against record icon templates.
2. Collect high-confidence rows where front and back agree as in-set templates.
3. Re-match all rows against the in-set templates.
4. Accept a final right_faction only when front and back agree after stage 2.
"""

from __future__ import annotations

import csv
import json
import math
import re
from datetime import datetime
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Font, PatternFill
from PIL import Image, ImageDraw, ImageFont, ImageOps


ROOT = Path(__file__).resolve().parents[1]
MAP_DATA = ROOT / "map" / "map-data.js"
ALLY_DIR = ROOT / "record" / "assets" / "ally"
OUT_DIR = ROOT / "tools" / "map-faction-right-auto"
CROP_DIR = OUT_DIR / "crops"
PAGES_DIR = OUT_DIR / "pages"
REVIEW_DIR = OUT_DIR / "review-pages"
OUT_CSV = OUT_DIR / "map-right-factions-auto.csv"
OUT_XLSX = OUT_DIR / "map-right-factions-auto.xlsx"
FONT_PATH = "C:/Windows/Fonts/msyh.ttc"
MASK_SIZE = 96
ACCEPT_MIN_CONFIDENCE = 0.55
ACCEPT_MIN_MARGIN = 0.08


FACTIONS_BY_CYCLE = {
    "c1": ["minoians", "labyrinthians", "hornsworn"],
    "c2": ["helots", "cyclopes", "symmachy"],
    "c3": ["sunheirs", "delphians", "twilightWatch"],
}

ASSET_BY_FACTION = {
    "cyclopes": "CYCLOPES.png",
    "delphians": "DELPHIANS.png",
    "helots": "HELOTS.png",
    "hornsworn": "HORNSWORN.png",
    "labyrinthians": "LABYRINTHIANS.png",
    "minoians": "MINOANS.png",
    "sunheirs": "SUNHEIRS.png",
    "symmachy": "SYMMACHY.png",
    "twilightWatch": "TWILIGHTWATCH.png",
}

SPECIAL_NO_FACTION = {("c3", "333"), ("c1", "T00")}

FIELDS = [
    "cycle",
    "tile_id",
    "label",
    "right_faction",
    "front_faction",
    "front_confidence",
    "front_margin",
    "back_faction",
    "back_confidence",
    "back_margin",
    "front_back_agree",
    "needs_review",
    "template_source",
    "preview_path",
    "note",
]


def load_tiles() -> list[dict[str, str]]:
    text = MAP_DATA.read_text(encoding="utf-8")
    match = re.search(r"window\.ATO_MAP_DATA\s*=\s*(\{.*\});?\s*$", text, re.S)
    if not match:
        raise RuntimeError("Could not parse map-data.js")
    data = json.loads(match.group(1))
    rows: list[dict[str, str]] = []
    for cycle in data["cycles"]:
        for tile in cycle["tiles"]:
            rows.append(
                {
                    "cycle": cycle["id"],
                    "tile_id": tile["id"],
                    "label": tile.get("label", tile["id"]),
                    "front": str((ROOT / "map" / tile["front"]).resolve()),
                    "back": str((ROOT / "map" / tile["back"]).resolve()),
                }
            )
    return rows


def front_context(image_path: str) -> Image.Image:
    image = Image.open(image_path).convert("RGB")
    width, height = image.size
    crop = image.crop((round(width * 0.50), round(height * 0.56), width, height))
    target_w = 520
    target_h = round(crop.height * (target_w / crop.width))
    return crop.resize((target_w, target_h), Image.Resampling.LANCZOS)


def front_outer_box(width: int, height: int) -> tuple[int, int, int, int]:
    return (
        round(width * 0.715),
        round(height * 0.655),
        round(width * 0.900),
        round(height * 0.865),
    )


def back_badge_box(image: Image.Image) -> tuple[int, int, int, int]:
    width, height = image.size
    return (
        round(width * 0.355),
        round(height * 0.038),
        round(width * 0.625),
        round(height * 0.158),
    )


def back_icon_box(image: Image.Image) -> tuple[int, int, int, int]:
    width, height = image.size
    return (
        round(width * 0.545),
        round(height * 0.068),
        round(width * 0.602),
        round(height * 0.134),
    )


def remove_small_components(mask: np.ndarray, min_area: int) -> np.ndarray:
    height, width = mask.shape
    seen = np.zeros(mask.shape, dtype=bool)
    out = np.zeros(mask.shape, dtype=bool)
    for start_y, start_x in zip(*np.nonzero(mask & ~seen)):
        stack = [(int(start_x), int(start_y))]
        seen[start_y, start_x] = True
        comp: list[tuple[int, int]] = []
        while stack:
            x, y = stack.pop()
            comp.append((x, y))
            for nx, ny in ((x - 1, y), (x + 1, y), (x, y - 1), (x, y + 1)):
                if 0 <= nx < width and 0 <= ny < height and mask[ny, nx] and not seen[ny, nx]:
                    seen[ny, nx] = True
                    stack.append((nx, ny))
        if len(comp) >= min_area:
            for x, y in comp:
                out[y, x] = True
    return out


def normalize_mask(mask: np.ndarray, size: int = MASK_SIZE) -> np.ndarray:
    ys, xs = np.nonzero(mask)
    if len(xs) == 0:
        return np.zeros((size, size), dtype=bool)
    x1, x2 = xs.min(), xs.max() + 1
    y1, y2 = ys.min(), ys.max() + 1
    cropped = Image.fromarray((mask[y1:y2, x1:x2] * 255).astype("uint8"), mode="L")
    canvas = Image.new("L", (size, size), 0)
    fitted = ImageOps.contain(cropped, (78, 78), method=Image.Resampling.LANCZOS)
    canvas.paste(fitted, ((size - fitted.width) // 2, (size - fitted.height) // 2))
    return np.asarray(canvas) > 96


def symbol_masks(slot: Image.Image, mode: str) -> list[np.ndarray]:
    slot = ImageOps.fit(slot.convert("RGB"), (MASK_SIZE, MASK_SIZE), method=Image.Resampling.LANCZOS, centering=(0.5, 0.5))
    arr = np.asarray(slot).astype(np.int16)
    gray = arr.mean(axis=2)
    r, g, b = arr[:, :, 0], arr[:, :, 1], arr[:, :, 2]
    saturation = arr.max(axis=2) - arr.min(axis=2)

    if mode == "front":
        candidates = [
            gray < np.percentile(gray, 30),
            (gray > np.percentile(gray, 72)) & (saturation < np.percentile(saturation, 68)),
        ]
        border = 10
    else:
        red_bg = (r > g + 16) & (r > b + 16) & (gray < 170)
        blue_bg = (b > r + 12) & (b > g + 8) & (gray < 170)
        brown_bg = (r > b + 12) & (g > b + 4) & (gray < 150)
        candidates = [(gray > np.percentile(gray, 78)) & ~(red_bg | blue_bg | brown_bg)]
        border = 5

    masks: list[np.ndarray] = []
    for mask in candidates:
        mask = mask.copy()
        mask[:border, :] = False
        mask[-border:, :] = False
        mask[:, :border] = False
        mask[:, -border:] = False
        cleaned = normalize_mask(remove_small_components(mask, min_area=10))
        if int(cleaned.sum()) >= 120:
            masks.append(cleaned)
    return masks


def asset_mask(faction: str) -> np.ndarray:
    image = Image.open(ALLY_DIR / ASSET_BY_FACTION[faction]).convert("RGBA")
    alpha = np.asarray(image)[:, :, 3]
    rgb = np.asarray(image.convert("RGB"))
    dark = rgb.mean(axis=2) < 180
    return normalize_mask(remove_small_components((alpha > 16) & dark, min_area=6))


def dice_score(a: np.ndarray, b: np.ndarray) -> float:
    denom = int(a.sum() + b.sum())
    if denom == 0:
        return 0.0
    return (2.0 * int((a & b).sum())) / denom


def score_mask(mask: np.ndarray, templates: list[np.ndarray]) -> float:
    best = 0.0
    for template in templates:
        best = max(best, dice_score(mask, template))
        for angle in (-6, -3, 3, 6):
            rotated = Image.fromarray((mask * 255).astype("uint8"), mode="L").rotate(angle, resample=Image.Resampling.BILINEAR)
            best = max(best, dice_score(np.asarray(rotated) > 96, template))
    return best


def match_slot(slot: Image.Image, mode: str, factions: list[str], templates: dict[str, list[np.ndarray]]) -> dict[str, object]:
    masks = symbol_masks(slot, mode)
    if not masks:
        return {"faction": "none", "confidence": 0.0, "margin": 0.0, "second": "", "mask": None}

    scores: list[tuple[str, float, np.ndarray]] = []
    for faction in factions:
        best_score = 0.0
        best_mask = masks[0]
        for mask in masks:
            score = score_mask(mask, templates[faction])
            if score > best_score:
                best_score = score
                best_mask = mask
        scores.append((faction, best_score, best_mask))
    scores.sort(key=lambda item: item[1], reverse=True)
    second_score = scores[1][1] if len(scores) > 1 else 0.0
    return {
        "faction": scores[0][0],
        "confidence": scores[0][1],
        "margin": scores[0][1] - second_score,
        "second": scores[1][0] if len(scores) > 1 else "",
        "mask": scores[0][2],
    }


def load_tile_masks(tile: dict[str, str]) -> tuple[Image.Image, Image.Image, Image.Image, Image.Image]:
    front = front_context(tile["front"])
    back = Image.open(tile["back"]).convert("RGB")
    front_slot = front.crop(front_outer_box(*front.size))
    back_slot = back.crop(back_icon_box(back))
    return front, back, front_slot, back_slot


def initial_rows(tiles: list[dict[str, str]], base_templates: dict[str, list[np.ndarray]]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for tile in tiles:
        front, back, front_slot, back_slot = load_tile_masks(tile)
        factions = FACTIONS_BY_CYCLE[tile["cycle"]]
        front_match = match_slot(front_slot, "front", factions, base_templates)
        back_match = match_slot(back_slot, "back", factions, base_templates)
        rows.append(
            {
                "tile": tile,
                "front": front,
                "back": back,
                "front_match": front_match,
                "back_match": back_match,
                "front_slot": front_slot,
                "back_slot": back_slot,
            }
        )
    return rows


def build_secondary_templates(rows: list[dict[str, object]], base_templates: dict[str, list[np.ndarray]]) -> dict[str, list[np.ndarray]]:
    templates = {faction: list(masks) for faction, masks in base_templates.items()}
    added: dict[str, int] = {faction: 0 for faction in templates}

    seeds = []
    for row in rows:
        tile = row["tile"]
        if (tile["cycle"], tile["tile_id"]) in SPECIAL_NO_FACTION:
            continue
        front_match = row["front_match"]
        back_match = row["back_match"]
        if front_match["faction"] != back_match["faction"]:
            continue
        if front_match["confidence"] < 0.58 or back_match["confidence"] < 0.50:
            continue
        if front_match["margin"] < 0.10 or back_match["margin"] < 0.10:
            continue
        seeds.append(row)

    seeds.sort(
        key=lambda row: min(float(row["front_match"]["confidence"]), float(row["back_match"]["confidence"])),
        reverse=True,
    )
    for row in seeds:
        faction = row["front_match"]["faction"]
        if added[faction] >= 8:
            continue
        front_masks = symbol_masks(row["front_slot"], "front")
        back_masks = symbol_masks(row["back_slot"], "back")
        if front_masks:
            templates[faction].append(front_masks[0])
        if back_masks:
            templates[faction].append(back_masks[0])
        added[faction] += 1
    return templates


def final_rows(tiles: list[dict[str, str]], templates: dict[str, list[np.ndarray]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for tile in tiles:
        front, back, front_slot, back_slot = load_tile_masks(tile)
        factions = FACTIONS_BY_CYCLE[tile["cycle"]]
        front_match = match_slot(front_slot, "front", factions, templates)
        back_match = match_slot(back_slot, "back", factions, templates)
        is_special = (tile["cycle"], tile["tile_id"]) in SPECIAL_NO_FACTION
        agree = front_match["faction"] == back_match["faction"] and front_match["faction"] != "none"
        strong = (
            float(front_match["confidence"]) >= ACCEPT_MIN_CONFIDENCE
            and float(back_match["confidence"]) >= ACCEPT_MIN_CONFIDENCE
            and float(front_match["margin"]) >= ACCEPT_MIN_MARGIN
            and float(back_match["margin"]) >= ACCEPT_MIN_MARGIN
        )
        accepted = agree and strong and not is_special
        note = "special_no_faction" if is_special else ""
        row = {
            "cycle": tile["cycle"],
            "tile_id": tile["tile_id"],
            "label": tile["label"],
            "right_faction": "none" if is_special else front_match["faction"] if accepted else "",
            "front_faction": str(front_match["faction"]),
            "front_confidence": f"{float(front_match['confidence']):.3f}",
            "front_margin": f"{float(front_match['margin']):.3f}",
            "back_faction": "none" if is_special else str(back_match["faction"]),
            "back_confidence": f"{float(back_match['confidence']):.3f}",
            "back_margin": f"{float(back_match['margin']):.3f}",
            "front_back_agree": "yes" if accepted else "",
            "needs_review": "" if accepted or is_special else "yes",
            "template_source": "secondary",
            "preview_path": "",
            "note": note,
        }
        row["preview_path"] = str(make_preview(tile, front, back, row).resolve())
        rows.append(row)
    return rows


def make_preview(tile: dict[str, str], front: Image.Image, back: Image.Image, row: dict[str, str]) -> Path:
    back_crop = back.crop(back_badge_box(back)).resize((520, 210), Image.Resampling.LANCZOS)
    preview = Image.new("RGB", (520, 550), "#f7f8f5")
    preview.paste(back_crop, (0, 0))
    preview.paste(front.resize((520, 245), Image.Resampling.LANCZOS), (0, 240))
    draw = ImageDraw.Draw(preview)
    small_font = ImageFont.truetype(FONT_PATH, 18)

    bx1, by1, bx2, by2 = back_icon_box(back)
    badge = back_badge_box(back)
    scale_x = 520 / (badge[2] - badge[0])
    scale_y = 210 / (badge[3] - badge[1])
    icon_box = (
        round((bx1 - badge[0]) * scale_x),
        round((by1 - badge[1]) * scale_y),
        round((bx2 - badge[0]) * scale_x),
        round((by2 - badge[1]) * scale_y),
    )
    draw.rectangle(icon_box, outline=(76, 175, 80), width=4)

    outer = front_outer_box(520, 245)
    draw.rectangle((outer[0], outer[1] + 240, outer[2], outer[3] + 240), outline=(239, 83, 80), width=5)

    draw.rectangle((0, 485, 519, 549), fill=(255, 255, 255))
    line1 = f"{row['cycle']}:{row['tile_id']} final={row['right_faction'] or 'review'}"
    line2 = f"F {row['front_faction']} {row['front_confidence']} m{row['front_margin']} | B {row['back_faction']} {row['back_confidence']} m{row['back_margin']}"
    draw.text((8, 491), line1, font=small_font, fill=(38, 50, 56))
    draw.text((8, 520), line2, font=small_font, fill=(84, 110, 122))

    path = CROP_DIR / f"{tile['cycle']}-{tile['tile_id']}.png"
    preview.save(path)
    return path


def write_csv(rows: list[dict[str, str]]) -> None:
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "right_factions_auto"
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    for col_idx, field in enumerate(FIELDS + ["preview"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(FIELDS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[field])
        img = XlsxImage(row["preview_path"])
        img.width = 260
        img.height = 275
        ws.add_image(img, f"P{row_idx}")
        ws.row_dimensions[row_idx].height = 205
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{len(rows) + 1}"
    widths = [9, 9, 10, 18, 18, 14, 12, 18, 14, 12, 16, 14, 16, 58, 22, 42]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.sheet_view.showGridLines = False
    try:
        wb.save(OUT_XLSX)
        return OUT_XLSX
    except OSError:
        fallback = OUT_DIR / f"map-right-factions-auto-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
        wb.save(fallback)
        return fallback


def write_contact_sheet(rows: list[dict[str, str]], path: Path, title: str, sort_review: bool = False) -> None:
    items = rows
    if sort_review:
        items = sorted(
            [row for row in rows if row["needs_review"] == "yes"],
            key=lambda row: (
                row["front_faction"] == row["back_faction"],
                min(float(row["front_margin"]), float(row["back_margin"])),
                min(float(row["front_confidence"]), float(row["back_confidence"])),
            ),
        )
    cols = 2
    cell_w, cell_h = 560, 620
    per_page = 6
    font = ImageFont.truetype(FONT_PATH, 22)
    page_dir = REVIEW_DIR if sort_review else PAGES_DIR
    page_dir.mkdir(parents=True, exist_ok=True)
    stem = "review" if sort_review else "auto"
    page_paths: list[Path] = []
    for page_index, start in enumerate(range(0, len(items), per_page), start=1):
        chunk = items[start : start + per_page]
        page_rows = math.ceil(len(chunk) / cols)
        sheet = Image.new("RGB", (cols * cell_w, page_rows * cell_h), "#f7f8f5")
        draw = ImageDraw.Draw(sheet)
        for idx, row in enumerate(chunk):
            x = (idx % cols) * cell_w
            y = (idx // cols) * cell_h
            preview = Image.open(row["preview_path"]).convert("RGB").resize((520, 550), Image.Resampling.LANCZOS)
            sheet.paste(preview, (x + 20, y + 12))
            draw.text((x + 22, y + 570), f"{row['cycle']}:{row['tile_id']} {row['right_faction'] or 'review'}", font=font, fill="#263238")
        page_path = page_dir / f"map-right-factions-{stem}-page-{page_index:02d}.png"
        sheet.save(page_path)
        page_paths.append(page_path)

    sections = "\n".join(
        f'<section><h2>Page {idx}</h2><img src="{page_dir.name}/{page.name}" alt="page {idx}"></section>'
        for idx, page in enumerate(page_paths, start=1)
    )
    path.write_text(
        f"""<!doctype html>
<html lang="zh-CN">
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title}</title>
<style>
body {{ margin:0; background:#f7f8f5; font-family:system-ui,sans-serif; color:#263238; }}
header {{ position:sticky; top:0; background:rgba(247,248,245,.96); padding:12px 14px; border-bottom:1px solid #d8dedf; }}
h1 {{ margin:0; font-size:20px; }}
h2 {{ margin:18px 12px 8px; font-size:16px; }}
img {{ display:block; width:100%; height:auto; }}
</style>
<header><h1>{title}</h1></header>
{sections}
</html>
""",
        encoding="utf-8",
    )


def main() -> int:
    CROP_DIR.mkdir(parents=True, exist_ok=True)
    PAGES_DIR.mkdir(parents=True, exist_ok=True)
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    tiles = load_tiles()
    base_templates = {faction: [asset_mask(faction)] for factions in FACTIONS_BY_CYCLE.values() for faction in factions}
    first_pass = initial_rows(tiles, base_templates)
    secondary_templates = build_secondary_templates(first_pass, base_templates)
    rows = final_rows(tiles, secondary_templates)
    write_csv(rows)
    xlsx_path = write_xlsx(rows)
    write_contact_sheet(rows, OUT_DIR / "index.html", "Map right faction auto tags")
    write_contact_sheet(rows, OUT_DIR / "review.html", "Map right faction review", sort_review=True)
    print(xlsx_path)
    print(OUT_CSV)
    print(OUT_DIR / "index.html")
    print(OUT_DIR / "review.html")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
